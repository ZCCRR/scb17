"""
====================================
author:ZCC
time:2020/11/22
E-mail:434445179@qq.com
====================================
"""
import openpyxl
import requests


def read_excel(file_name, sh_name):
    """
    读取Excel数据
    :param sh_name: 表单名
    :param file_name:  文件名
    :return: Excel文件数据（列表）
    """
    # 找到工作簿（Excel）
    wb = openpyxl.load_workbook(file_name)  # 打开文件
    sh = wb[sh_name]  # 选中表单
    max_row = sh.max_row  # 获取表单中的最大的行数
    li = []  # 定义一个空列表
    for i in range(2, max_row + 1):  # 取头不取尾 左闭右开
        dict_1 = dict(
            id=sh.cell(row=i, column=1).value,  # 取出ID
            url=sh.cell(row=i, column=5).value,  # 取出url
            data=sh.cell(row=i, column=6).value,  # 取出data
            expect=sh.cell(row=i, column=7).value)  # 取出expect
        li.append(dict_1)  # 把所有的数据都添加到空列表中
    return li


# print(read_excel("test_case_api.xlsx","register"))

def api_func(url, data):
    """
    发送请求
    :param url: url地址
    :param data: data数据
    :return: 响应数据
    """
    header = {"X-Lemonban-Media-Type": "lemonban.v2", "Content-Type": "application/json"}
    res = requests.post(url=url, json=data, headers=header)
    response = res.json()
    return response


def write_excel(file_name, sh_name, row, column,result):
    """
    写入断言结果
    :param file_name: 文件名
    :param sh_name: 表单名
    :param row: 行
    :param column:  列
    :return:
    """
    wb = openpyxl.load_workbook(file_name)  # 打开Excel文件
    sh = wb[sh_name]  # 读取表单
    result = sh.cell(row=row, column=column).value  # 获取表单数据
    wb.save(file_name)  # 将结果写入Excel文件


def func(file_name,sh_name):
    cases = read_excel(file_name, sh_name)
    for case in cases:
        id = case.get("id")  # 取出ID
        url = case.get("url")  # 取出url
        data = case.get("data")  # 取出data
        expect = case.get("expect")  # 取出预期结果
        # print(type(expect)) # 查看except的类型
        # expect = eval(expect)  # 通过eval识别字符串中的表达式 转换为字典格式
        # print(type(expect))
        expect_msg = eval(expect).get('msg')  # 取出预期结果中的msg信息
        # 获取实际结果
        result = api_func(url=url, data=eval(data))
        # 获取实际结果中的msg信息
        result_msg = result.get('msg')
        print("实际结果为：{}".format(result_msg))
        print("预期结果为：{}".format(expect_msg))
        if expect_msg== result_msg:
            print("该条测试用例执行通过！")
            final_res = "通过"
        else:
            print('这条测试用例不通过！！！')
            final_res = "不通过"
        print("*"*30)
        write_excel(file_name,sh_name,id+1,8,final_res)
func("test_case_api.xlsx","register")



